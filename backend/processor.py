# /backend/processor.py
import os
import glob
import time
import random
import copy
import pandas as pd
import numpy as np
import openpyxl
from datetime import datetime
from collections import defaultdict
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.formula.translate import Translator
try:
    from openpyxl.formula.array import ArrayFormula
except ImportError:
    ArrayFormula = None

def process_rider_data(city, selected_option, source_folder, base_path, log_callback, progress_callback,
                       finish_callback, theme, workspace_path=None):
    def log(msg, level="INFO"):
        prefix = {"INFO": "[INFO]", "WARN": "[WARN]", "ERROR": "[ERRO]", "SYSTEM": "[SYS ]", "SUCCESS": "[ OK ]"}.get(
            level, "[INFO]")
        log_callback(f"{prefix} {msg}\n", level)

    def clean_wb_str(val):
        if pd.isna(val) or val is None or str(val).strip().lower() == 'nan':
            return ""
        s = str(val).strip().lstrip("'")
        if s.endswith('.0'):
            s = s[:-2]
        return s

    def to_numeric_if_possible(val_str):
        if not val_str:
            return val_str
        try:
            f = float(val_str)
            if f.is_integer():
                return int(f)
            return f
        except ValueError:
            return val_str

    def _norm_date(val):
        if pd.isna(val) or val is None: return ""
        s = str(val).strip().split(' ')[0].replace('/', '-')
        try:
            parts = s.split('-')
            if len(parts) >= 3:
                return f"{parts[0]}-{int(parts[1]):02d}-{int(parts[2]):02d}"
        except:
            pass
        return s

    def get_id_name(row_series):
        r_id, r_name = "", ""
        for c in row_series.index:
            c_str = str(c)
            c_lower = c_str.lower()
            if ('id' in c_lower or '编号' in c_lower) and not any(
                    x in c_lower for x in ['运单', '订单', '站', '团队', '门店', '网点', '商户']):
                r_id = str(row_series[c])
            if c_str in ['骑手姓名', '配送员姓名', '姓名', '员工姓名', '骑手名称'] or \
                    ('姓名' in c_str and '站' not in c_str) or \
                    ('名称' in c_str and '骑手' in c_str):
                r_name = str(row_series[c])
        return r_id.replace('.0', '').strip(), r_name.replace('.0', '').strip()

    stats_info = {
        "riders": 0, "orders": 0, "penalty_orders": 0, "problem_orders": 0, "elapsed_time": 0.0, "out_folder": ""
    }

    try:
        start_time = time.time()
        progress_callback(0.02, "召唤系统核心中...")
        log(f">>> {random.choice(theme['msg_start'])}", "SYSTEM")

        log(f">>> {random.choice(theme['msg_awake'])}", "SYSTEM")
        progress_callback(0.05, "偷瞄极品草料名单...")
        log(f"锁定目标区: [{city}], 任务节奏: [{selected_option}]")

        keywords = ["欺诈单", "问题单", "违规", "配送费", "骑手支付绑定", "兼职价格档案"]
        ar = ["是否剔除", "剔除原因", "扣款金额"]

        dict1 = {
            "配送费": [5, 0, 13], "问题单": [3, 1, 2, 9, 10, 11], "违规扣款": [5, 0, 13, 17, 18, 19]
        }

        valid_riders_dates = set()
        dictp_fraud = set()
        dict2_late = set()
        price_mapping_id = {}
        price_mapping_name = {}

        if os.path.isfile(base_path):
            main_wb_path = base_path
            base_dir = os.path.dirname(base_path)
        else:
            main_wb_path = os.path.join(base_path, "config.xlsx")
            base_dir = base_path

        if not os.path.exists(main_wb_path): raise FileNotFoundError(f"哎呀！配置文件《config.xlsx》不小心弄丢啦: {main_wb_path}")

        df_apply = pd.read_excel(main_wb_path, sheet_name="申请名单", dtype={1: str})
        df_rules = pd.read_excel(main_wb_path, sheet_name="扣款标准")

        team_mapping = {}
        try:
            df_team = pd.read_excel(main_wb_path, sheet_name="团队名称", dtype=str)
            name_col = next((c for c in df_team.columns if "名称" in str(c)), df_team.columns[0])
            id_col = next((c for c in df_team.columns if "id" in str(c).lower() or "编码" in str(c)),
                          df_team.columns[1] if len(df_team.columns) > 1 else None)
            if id_col is not None:
                team_mapping = dict(zip(df_team[name_col].fillna("").astype(str).str.strip(),
                                        df_team[id_col].fillna("").astype(str).str.strip()))
            log(f"成功加载【团队名称】映射表，共备好 {len(team_mapping)} 条匹配锦囊~")
        except Exception as e:
            log(f"未找到【团队名称】表或解析失败 (配送所得表A列将留空): {e}", "WARN")

        first_date = pd.to_datetime(df_apply.iloc[:, 3]).min()
        last_date = pd.to_datetime(df_apply.iloc[:, 4]).max()
        df_apply.iloc[:, 3] = pd.to_datetime(df_apply.iloc[:, 3])
        df_apply.iloc[:, 4] = pd.to_datetime(df_apply.iloc[:, 4])

        for row in df_apply.itertuples(index=False):
            try:
                rider_id = str(row[1]).replace('.0', '').strip()
                date_range = pd.date_range(row[3], row[4])
                for d in date_range: valid_riders_dates.add(f"{rider_id}|{d.strftime('%Y-%m-%d')}")
            except:
                continue

        apply_riders_count = len(df_apply.iloc[:, 1].dropna().unique())
        log(f"名单解读成功，发现排班表中登记了 {apply_riders_count} 个打工人（共对应 {len(valid_riders_dates)} 个出勤日次）~")
        progress_callback(0.15, "名单确认完毕")

        output_base = workspace_path if workspace_path else base_dir
        out_folder = os.path.join(output_base, f"{last_date.month}月{city}兼职核算")
        stats_info["out_folder"] = out_folder
        os.makedirs(out_folder, exist_ok=True)

        log(">>> 正在扫描历史文件库，启动防重溯源雷达...", "SYSTEM")
        existing_waybills = {"配送单": {}, "违规索赔": {}, "问题单": {}}
        if os.path.exists(out_folder):
            for file in glob.glob(os.path.join(out_folder, "*.xlsx")):
                if "(重复)" in os.path.basename(file) or "~$" in os.path.basename(file): continue
                try:
                    file_basename = os.path.basename(file)
                    for sht in ["配送单", "违规索赔", "问题单"]:
                        try:
                            df_hist = pd.read_excel(file, sheet_name=sht, dtype=str)
                            headers = df_hist.columns.tolist()
                            if sht == "问题单":
                                wb_idx = next((c for c in headers if any(kw in str(c) for kw in ["运单", "订单", "单号"])), None)
                                type_idx = headers[0] if len(headers) > 0 else None
                                if wb_idx and type_idx:
                                    for t, w in zip(df_hist[type_idx], df_hist[wb_idx]):
                                        w_cl = clean_wb_str(w)
                                        if w_cl: existing_waybills[sht][f"{t}_{w_cl}"] = file_basename
                            else:
                                fallback = headers[10] if (sht == "违规索赔" and len(headers)>10) else None
                                wb_idx = next((c for c in headers if any(kw in str(c) for kw in ["运单", "订单", "单号"])), fallback)
                                if wb_idx:
                                    for w in df_hist[wb_idx]:
                                        w_cl = clean_wb_str(w)
                                        if w_cl: existing_waybills[sht][w_cl] = file_basename
                        except Exception:
                            pass
                except Exception as e:
                    log(f"读取历史记录防重失败: {os.path.basename(file)} - {str(e)}", "WARN")

        intercept_records = []

        project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        
        # 尝试多个可能存放 template 的目录:
        # 1. 工程同级目录 (project_root 的父级)
        # 2. 工程根目录 (project_root)
        # 3. 配置文件同级目录 (base_dir)
        possible_template_dirs = [
            os.path.join(os.path.dirname(project_root), "template"),
            os.path.join(project_root, "template"),
            os.path.join(base_dir, "template")
        ]
        
        template_path = ""
        for t_dir in possible_template_dirs:
            potential_path = os.path.join(t_dir, f"{city}-template.xlsx")
            if os.path.exists(potential_path):
                template_path = potential_path
                break
                
        if not template_path:
            raise FileNotFoundError(f"哎呀！运行过程中出了个小意外：模板文件找不到了 (需包含 {city}-template.xlsx)。请确保 template 目录存放在工程同级目录下。")

        log(f"开始疯狂填表: {os.path.basename(template_path)}", "SYSTEM")
        wb1 = openpyxl.load_workbook(template_path)
        
        def fast_recreate_sheet(wb, sheet_name):
            if sheet_name in wb.sheetnames:
                idx = wb.sheetnames.index(sheet_name)
                old_ws = wb[sheet_name]
                tab_color = old_ws.sheet_properties.tabColor
                del wb[sheet_name]
                new_ws = wb.create_sheet(title=sheet_name, index=idx)
                if tab_color:
                    try:
                        new_ws.sheet_properties.tabColor = copy.copy(tab_color)
                    except:
                        pass
                return new_ws
            return wb.create_sheet(title=sheet_name)
            
        ws_apply = fast_recreate_sheet(wb1, "申请名单")

        headers_apply = list(df_apply.columns)
        ws_apply.append(headers_apply)
        for r_idx, row in enumerate(df_apply.itertuples(index=False), 2):
            ws_apply.append(list(row))
            for c_idx in [4, 5]:
                if c_idx <= len(row): ws_apply.cell(row=r_idx, column=c_idx).number_format = 'yyyy/mm/dd'

        progress_callback(0.25, "基础模板就绪")

        log(">>> 正在库里寻觅待处理源文件...", "SYSTEM")
        files_to_process = []
        for kw in keywords:
            search_pattern = os.path.join(source_folder, f"*{kw}*.*")
            for f in glob.glob(search_pattern):
                if f.lower().endswith(('.xls', '.xlsx', '.csv')): files_to_process.append((f, kw))

        total_files = len(files_to_process)
        log(f"寻觅结束，抱回了 {total_files} 个文件袋")
        if total_files == 0: log(random.choice(theme['msg_empty']), "WARN")

        processed_count = 0
        current_progress = 0.25
        progress_step = 0.6 / total_files if total_files > 0 else 0

        for file_path, kw in files_to_process:
            wb_name = os.path.basename(file_path)
            log(random.choice(theme['msg_process']).format(wb_name=wb_name))

            if file_path.lower().endswith('.csv'):
                try:
                    df_source = pd.read_csv(file_path, encoding='gbk', dtype=str, low_memory=False)
                    if len(df_source.columns) <= 2: df_source = pd.read_csv(file_path, encoding='gbk', sep='\t',
                                                                            dtype=str)
                except UnicodeDecodeError:
                    df_source = pd.read_csv(file_path, encoding='utf-8', dtype=str, low_memory=False)
                    if len(df_source.columns) <= 2: df_source = pd.read_csv(file_path, encoding='utf-8', sep='\t',
                                                                            dtype=str)
            else:
                df_source = pd.read_excel(file_path, dtype=str)

            for col in df_source.columns: 
                df_source[col] = df_source[col].fillna("").astype(str).str.strip(' ="\t\r\n')
                # 如果可能，尝试将列重新转换为数字（忽略空字符串）
                try:
                    s_replaced = df_source[col].copy()
                    s_replaced.replace("", np.nan, inplace=True)
                    s_num = pd.to_numeric(s_replaced)
                    # 如果转换成功，且没有将有效的字符串强制转换为 NaN（除非它们是空的）
                    if not s_num.isna().any() or s_replaced.isna().equals(s_num.isna()):
                        df_source[col] = s_num
                except Exception:
                    pass

            if "兼职价格档案" in wb_name:
                try:
                    if not file_path.lower().endswith('.csv'):
                        df_price = pd.read_excel(file_path, sheet_name="兼职价格档案明细", dtype=str)
                        for col in df_price.columns:
                            df_price[col] = df_price[col].fillna("").astype(str).str.strip(' ="\t\r\n')
                    else:
                        df_price = df_source.copy()
                except Exception as e:
                    log(f"未能找到【兼职价格档案明细】工作表，尝试从默认表读取", "WARN")
                    df_price = df_source.copy()

                rider_id_idx = next((i for i, c in enumerate(df_price.columns) if
                                     ('id' in str(c).lower() or '编号' in str(c)) and ('骑手' in str(c) or '员工' in str(c))),
                                    None)
                rider_name_idx = next((i for i, c in enumerate(df_price.columns) if
                                       ('名' in str(c) or '姓名' in str(c)) and ('骑手' in str(c) or '员工' in str(c))), None)
                if rider_id_idx is None: rider_id_idx = 2
                if rider_name_idx is None: rider_name_idx = 3

                for r in df_price.itertuples(index=False):
                    try:
                        if len(r) > 8:
                            date_val = _norm_date(r[1])
                            station_val = str(r[5]).strip()
                            price_val = str(r[8]).strip()
                            r_id = str(r[rider_id_idx]).replace('.0', '').strip() if rider_id_idx < len(r) else ""
                            r_name = str(r[rider_name_idx]).strip() if rider_name_idx < len(r) else ""
                            if r_id: price_mapping_id[f"{date_val}_{station_val}_{r_id}"] = price_val
                            if r_name: price_mapping_name[f"{date_val}_{station_val}_{r_name}"] = price_val
                    except Exception:
                        pass

                log(f"-> 🎯 蓝橙单价档案装载成功！(ID规则数: {len(price_mapping_id)} |姓名规则数: {len(price_mapping_name)})", "INFO")
                processed_count += 1
                current_progress += progress_step
                progress_callback(current_progress, f"狂奔中 {processed_count}/{total_files}: {wb_name[:12]}...")
                continue

            match_key = next((k for k in dict1.keys() if k in wb_name), None)
            if match_key:
                cols = dict1[match_key]
                rider_col, date_col, wb_col = cols[0], cols[1], cols[2]
                max_needed_col = max(cols)
                while len(df_source.columns) <= max_needed_col: df_source[f"Temp_Col_{len(df_source.columns)}"] = np.nan
                waybill_col_name = df_source.columns[wb_col]
                df_source[waybill_col_name] = df_source[waybill_col_name].astype(str).str.lstrip("'")

                try:
                    rider_series = df_source.iloc[:, rider_col].astype(str).str.replace('.0', '',
                                                                                        regex=False).str.strip()
                    date_series = pd.to_datetime(df_source.iloc[:, date_col].astype(str).str.split(' ').str[0],
                                                 errors='coerce').dt.strftime('%Y-%m-%d')
                    combined_series = rider_series + "|" + date_series
                    mask_valid = combined_series.isin(valid_riders_dates)
                    df_source = df_source[mask_valid].copy()
                except Exception as e:
                    log(f"哎呀，过滤数据时摔了一跤: {str(e)}", "WARN")

                if "问题单" in wb_name:
                    mask_late = df_source.iloc[:, 0].astype(str).str.contains("不准时单", na=False)
                    dict2_late.update(df_source.loc[mask_late, df_source.columns[wb_col]].tolist())

                if "配送费" in wb_name:
                    offset = 1 if selected_option == "全职" else 0
                    target_max = 22 + offset
                    while len(df_source.columns) <= target_max: df_source[f"Temp_Col_{len(df_source.columns)}"] = np.nan
                    col1 = df_source.columns[19 + offset]
                    col2 = df_source.columns[20 + offset]
                    col3 = df_source.columns[21 + offset]
                    df_source[col1] = df_source.iloc[:, wb_col].apply(lambda x: "完成单-不准时" if x in dict2_late else "完成单")
                    df_source[col2] = pd.to_datetime(df_source.iloc[:, date_col]).dt.dayofweek.apply(
                        lambda x: "是" if x in [5, 6] else "否")
                    df_source[col3] = df_source.iloc[:, wb_col].apply(lambda x: "是" if x in dictp_fraud else "否")

                if match_key != "配送费":
                    col_names = list(df_source.columns)
                    for i in range(3, 6): col_names[cols[i]] = ar[i - 3]
                    df_source.columns = col_names

            if "欺诈单" in wb_name:
                if "运单id" in df_source.columns: dictp_fraud.update(df_source["运单id"].astype(str).tolist())

            elif "配送费" in wb_name:
                ws_sht0 = wb1["配送单"]
                ws_sht1 = wb1["配送所得表"]
                ws_sht2 = wb1["安全基金"]
                cols_to_drop = [1, 2, 3, 7]
                if selected_option == "全职":
                    cols_to_drop.extend(list(range(23, 38)))
                else:
                    cols_to_drop.extend(list(range(22, 38)))
                cols_to_drop = [c for c in cols_to_drop if c < len(df_source.columns)]
                df_source.drop(df_source.columns[cols_to_drop], axis=1, inplace=True)
                df_source = df_source[~df_source.iloc[:, 6].astype(str).str.contains("异常", na=False)]

                wb_col_name = next((c for c in df_source.columns if any(kw in str(c) for kw in ["运单", "订单", "单号"])),
                                   None)
                if wb_col_name:
                    clean_waybills = df_source[wb_col_name].apply(clean_wb_str)
                    mask_dup = clean_waybills.isin(existing_waybills["配送单"].keys())
                    df_dups = df_source[mask_dup].copy()
                    df_source = df_source[~mask_dup].copy()
                    stats_info["orders"] += len(df_source)
                    if not df_dups.empty:
                        dup_sources = {}
                        for row in df_dups.itertuples(index=False):
                            wb_val = clean_wb_str(row[wb_col_name])
                            src_file = existing_waybills["配送单"].get(wb_val)
                            if src_file:
                                r_id = str(row[2]).replace('.0', '').strip() if len(row) > 2 else ""
                                r_name = str(row[3]).strip() if len(row) > 3 else ""
                                intercept_records.append((src_file, "配送费", wb_val, r_id, r_name))
                                dup_sources[src_file] = dup_sources.get(src_file, 0) + 1
                        for src_file, count in dup_sources.items(): log(
                            f"💥 [红牌拦截] 发现 {count} 条【配送单】重复记录！(源自历史文件: {src_file})", "ERROR")

                for col in df_source.columns:
                    sample = str(df_source[col].iloc[0]) if not df_source[col].empty else ""
                    if len(sample) < 15:
                        try:
                            df_source[col] = pd.to_numeric(df_source[col])
                        except Exception:
                            pass

                headers_sht0 = list(df_source.columns)
                ws_sht0 = fast_recreate_sheet(wb1, "配送单")
                ws_sht0.append(headers_sht0)
                for row_data in df_source.values.tolist():
                    ws_sht0.append(row_data)
                
                last_col = 17 if selected_option == "全职" else 16
                headers = ["运单状态", "是否周末", "是否欺诈单"]
                for i, h in enumerate(headers): ws_sht0.cell(row=1, column=last_col + i, value=h)

                if len(df_source) > 0:
                    df_sht2 = df_source.iloc[:, 1:4].copy()
                else:
                    df_sht2 = pd.DataFrame(columns=["团队名称", "骑手ID", "骑手姓名"])

                # 确保申请名单中的所有骑手都被匹配到工作簿中
                try:
                    df_apply_local = df_apply.copy()
                    if len(df_apply_local.columns) > 0 and city in df_apply_local.iloc[:, 0].astype(str).unique():
                        df_apply_local = df_apply_local[df_apply_local.iloc[:, 0].astype(str) == city]
                        
                    df_apply_riders = pd.DataFrame()
                    cols = df_apply_local.columns
                    
                    team_col = next((c for c in cols if '团队' in str(c) or '站' in str(c) or '网点' in str(c)), None)
                    id_col = next((c for c in cols if 'id' in str(c).lower() or '编号' in str(c)), cols[1] if len(cols) > 1 else None)
                    name_col = next((c for c in cols if '名' in str(c)), cols[2] if len(cols) > 2 else None)
                    
                    team_ser = df_apply_local[team_col] if team_col else pd.Series([""] * len(df_apply_local))
                    id_ser = df_apply_local[id_col].astype(str).str.replace('.0', '', regex=False).str.strip() if id_col else pd.Series([""] * len(df_apply_local))
                    name_ser = df_apply_local[name_col] if name_col else pd.Series([""] * len(df_apply_local))
                    
                    df_apply_riders = pd.DataFrame({"团队名称": team_ser.values, "骑手ID": id_ser.values, "骑手姓名": name_ser.values})
                    df_apply_riders.columns = df_sht2.columns if len(df_sht2.columns) == 3 else ["团队名称", "骑手ID", "骑手姓名"]
                    
                    df_sht2 = pd.concat([df_sht2, df_apply_riders], ignore_index=True)
                except Exception as e:
                    log(f"合并申请名单数据失败: {e}", "WARN")

                df_sht2.replace(r'^\s*$', np.nan, regex=True, inplace=True)
                if len(df_sht2.columns) > 1:
                    df_sht2.iloc[:, 1] = df_sht2.iloc[:, 1].astype(str).str.replace('.0', '', regex=False).str.strip()
                    df_sht2.iloc[:, 1] = df_sht2.iloc[:, 1].replace(r'^\s*$', np.nan, regex=True).replace("nan", np.nan).replace("None", np.nan)
                    df_sht2.iloc[:, 1] = df_sht2.iloc[:, 1].replace("", np.nan)
                    df_sht2.dropna(subset=[df_sht2.columns[1]], inplace=True)
                    # 重新转换为数字以在输出 Excel 中保留数字类型
                    try:
                        df_sht2.iloc[:, 1] = pd.to_numeric(df_sht2.iloc[:, 1])
                    except:
                        pass
                df_sht2.drop_duplicates(subset=[df_sht2.columns[1]], keep='first', inplace=True)
                df_sht2.sort_values(by=df_sht2.columns[0], ascending=True, inplace=True)
                df_sht2.reset_index(drop=True, inplace=True)
                df_sht2.fillna("", inplace=True)

                data_list = df_sht2.values.tolist()
                headers_sht2 = df_sht2.columns.tolist() if len(df_sht2.columns) >= 3 else ["团队名称", "骑手ID", "骑手姓名"]

                for c_idx, val in enumerate(headers_sht2, 1): ws_sht2.cell(row=1, column=c_idx, value=val)
                for r_idx, row_data in enumerate(data_list, 2):
                    for c_idx, val in enumerate(row_data, 1): ws_sht2.cell(row=r_idx, column=c_idx, value=val)

                last_row_sht2 = len(data_list) + 1
                num_riders = last_row_sht2 - 1
                stats_info["riders"] = max(stats_info.get("riders", 0), num_riders)
                log(f"-> 🎯 统计完毕：【安全基金】结算骑手共计 {num_riders} 人，总计 {stats_info['orders']} 单！", "INFO")

                if last_row_sht2 > 2:
                    translator_cache_sht2 = {}
                    for col_idx in range(4, ws_sht2.max_column + 1):
                        cell_template = ws_sht2.cell(row=2, column=col_idx)
                        if cell_template.value and isinstance(cell_template.value,
                                                              str) and cell_template.value.startswith('='):
                            clean_f = cell_template.value
                            if clean_f not in translator_cache_sht2:
                                translator_cache_sht2[clean_f] = Translator(clean_f, origin=cell_template.coordinate)
                            translator = translator_cache_sht2[clean_f]
                            for row_idx in range(3, last_row_sht2 + 1):
                                target_cell = ws_sht2.cell(row=row_idx, column=col_idx)
                                try:
                                    translated_f = translator.translate_formula(target_cell.coordinate)
                                    target_cell.value = translated_f
                                except Exception:
                                    target_cell.value = cell_template.value
                                if cell_template.has_style: target_cell._style = copy.copy(cell_template._style)

                if num_riders > 0:

                    m_ranges = list(ws_sht1.merged_cells.ranges)
                    for m_range in m_ranges:
                        if m_range.min_row >= 4: ws_sht1.unmerge_cells(m_range.coord)

                    if num_riders > 1:
                        try:
                            ws_sht1.insert_rows(5, amount=num_riders - 1)
                        except:
                            pass

                    max_col_sht1 = ws_sht1.max_column
                    template_row_info = []
                    for c_idx in range(1, max_col_sht1 + 1):
                        cell = ws_sht1.cell(row=4, column=c_idx)
                        template_row_info.append({
                            'val': cell.value, 'style': copy.copy(cell._style) if cell.has_style else None,
                            'col_letter': get_column_letter(c_idx)
                        })

                    translator_cache_sht1 = {}
                    for i in range(num_riders):
                        t_row = i + 4
                        for c_idx, t_item in enumerate(template_row_info, 1):
                            target_cell = ws_sht1.cell(row=t_row, column=c_idx)
                            if t_row > 4 and t_item['style']: target_cell._style = t_item['style']

                            if c_idx in [2, 3, 4]:
                                val_idx = c_idx - 2
                                target_cell.value = ws_sht2.cell(row=i + 2, column=val_idx + 1).value
                            elif c_idx == 1:
                                team_name = str(ws_sht2.cell(row=i + 2, column=1).value or "").strip()
                                t_id = team_mapping.get(team_name, "")
                                target_cell.value = int(t_id) if t_id and str(t_id).isdigit() else t_id
                            else:
                                formula_val = t_item['val']
                                if isinstance(formula_val, str) and formula_val.startswith('='):
                                    clean_f = formula_val.replace("{", "").replace("}", "").strip()
                                    if clean_f not in translator_cache_sht1:
                                        origin_coord = t_item['col_letter'] + "4"
                                        translator_cache_sht1[clean_f] = Translator(clean_f, origin=origin_coord)
                                    translator = translator_cache_sht1[clean_f]
                                    try:
                                        translated_f = translator.translate_formula(target_cell.coordinate)
                                        if (c_idx == 1 or c_idx == 20) and ArrayFormula:
                                            target_cell.value = ArrayFormula(target_cell.coordinate, translated_f)
                                        else:
                                            target_cell.value = translated_f
                                    except:
                                        target_cell.value = clean_f
                                else:
                                    if t_row > 4: target_cell.value = t_item['val']

                    f_row = ws_sht1.max_row
                    for c_clear in range(2, 5): ws_sht1.cell(row=f_row, column=c_clear).value = None
                    try:
                        ws_sht1.merge_cells(f"A{f_row}:D{f_row}")
                    except:
                        pass

            elif "违规" in wb_name:
                cols_to_drop = [1, 2, 3, 15, 16]
                cols_to_drop = [c for c in cols_to_drop if c < len(df_source.columns)]
                df_source.drop(df_source.columns[cols_to_drop], axis=1, inplace=True)

                data_arr = df_source.values.tolist()
                headers = df_source.columns.tolist()

                for row in data_arr:
                    while len(row) < 15: row.append("")
                    if "泉州" in str(row[1]) and "索赔" in str(row[6]):
                        if "提前点送达" in str(row[7]):
                            row[14] = float(row[4] or 0) - 50
                        elif "配送超时" in str(row[7]) or "未进行配送" in str(row[7]):
                            row[14] = float(row[4] or 0)
                        else:
                            row[14] = float(row[4] or 0) - 20
                    else:
                        row[14] = float(row[4] or 0)
                    if "欺诈单" in str(row[6]): row[12], row[13], row[14] = "是", "欺诈单不考核", 0

                ws_wg = wb1["违规索赔"]
                start_row = 1
                while ws_wg.cell(row=start_row, column=1).value is not None: start_row += 1

                if start_row == 1:
                    for c_idx, h_val in enumerate(headers, 1): ws_wg.cell(row=start_row, column=c_idx, value=h_val)
                    start_row += 1

                added_penalty_count = 0
                wb_idx = next((i for i, c in enumerate(headers) if any(kw in str(c) for kw in ["运单", "订单", "单号"])), 10)

                id_idx, name_idx = -1, -1
                for i, c in enumerate(headers):
                    c_str = str(c)
                    c_lower = c_str.lower()
                    if ('id' in c_lower or '编号' in c_lower) and not any(
                            x in c_lower for x in ['运单', '订单', '站', '团队']): id_idx = i
                    if c_str in ['骑手姓名', '配送员姓名', '姓名', '员工姓名', '骑手名称'] or ('姓名' in c_str and '站' not in c_str) or (
                            '名称' in c_str and '骑手' in c_str): name_idx = i

                if wb_idx is not None and wb_idx < len(headers):
                    dup_sources = {}
                    for row in data_arr:
                        wb_val = clean_wb_str(row[wb_idx]) if wb_idx < len(row) else ""
                        if wb_val and wb_val != "-" and wb_val in existing_waybills["违规索赔"]:
                            src_file = existing_waybills["违规索赔"][wb_val]
                            r_id = str(row[id_idx]).replace('.0', '').strip() if id_idx != -1 else ""
                            r_name = str(row[name_idx]).strip() if name_idx != -1 else ""
                            intercept_records.append((src_file, "索赔扣款", wb_val, r_id, r_name))
                            dup_sources[src_file] = dup_sources.get(src_file, 0) + 1
                        else:
                            ws_wg.append(list(row))
                            added_penalty_count += 1
                    for src_file, count in dup_sources.items(): log(
                        f"💥 [红牌拦截] 发现 {count} 条【违规索赔】重复记录！(源自历史文件: {src_file})", "ERROR")
                else:
                    for row in data_arr:
                        ws_wg.append(list(row))
                        added_penalty_count += 1

                stats_info["penalty_orders"] += added_penalty_count

            elif "问题单" in wb_name:
                while len(df_source.columns) < 13: df_source[f"Temp_Col_{len(df_source.columns)}"] = ""
                col_names = list(df_source.columns)
                col_names[12] = "是否欺诈单"
                df_source.columns = col_names

                cols_to_drop = [c for c in df_source.columns if "Unnamed" in str(c)]
                if cols_to_drop: df_source.drop(columns=cols_to_drop, inplace=True)

                try:
                    df_source.iloc[:, 1] = pd.to_datetime(df_source.iloc[:, 1].astype(str).str.split(' ').str[0],
                                                          errors='coerce').dt.strftime('%Y-%m-%d')
                    df_source.iloc[:, 1] = df_source.iloc[:, 1].fillna("")
                except Exception:
                    pass

                rules_headers = df_rules.columns.tolist()
                rules_data = df_rules.values.tolist()

                is_fraud_info = df_source.iloc[:, 2].astype(str).isin(dictp_fraud).map({True: "是", False: "否"}).tolist()
                col0_list = df_source.iloc[:, 0].tolist()
                col5_list = df_source.iloc[:, 5].astype(str).tolist()
                
                fast_vals = []
                for r0, r5, fraud in zip(col0_list, col5_list, is_fraud_info):
                    val = np.nan
                    if fraud == "是":
                        val = 0
                    elif r0 in rules_headers:
                        wtd_idx = rules_headers.index(r0)
                        for rule_row in rules_data:
                            if str(rule_row[1]) in r5:
                                val = rule_row[wtd_idx]
                                break
                    fast_vals.append(val)
                
                df_source.iloc[:, 12] = is_fraud_info
                df_source.iloc[:, 11] = fast_vals

                ws_wtd = wb1["问题单"]
                clean_problem_waybills = df_source.iloc[:, 2].apply(clean_wb_str)
                combined_keys = df_source.iloc[:, 0].astype(str).str.strip() + "_" + clean_problem_waybills
                mask_dup = combined_keys.isin(existing_waybills["问题单"].keys())

                df_dups = df_source[mask_dup].copy()
                df_source = df_source[~mask_dup].copy()
                stats_info["problem_orders"] += len(df_source)

                if not df_dups.empty:
                    dup_sources = {}
                    for row in df_dups.itertuples(index=False):
                        col0_val = str(row[0]).strip() if len(row) > 0 else ""
                        wb_val = clean_wb_str(row[2]) if len(row) > 2 else ""
                        key = f"{col0_val}_{wb_val}"
                        src_file = existing_waybills["问题单"].get(key)
                        if src_file:
                            val_type = col0_val if col0_val else "未知问题"
                            r_id = str(row[2]).replace('.0', '').strip() if len(row) > 2 else ""
                            r_name = str(row[3]).strip() if len(row) > 3 else ""
                            intercept_records.append((src_file, val_type, wb_val, r_id, r_name))
                            dup_sources[src_file] = dup_sources.get(src_file, 0) + 1
                    for src_file, count in dup_sources.items(): log(
                        f"💥 [红牌拦截] 发现 {count} 条【问题单】重复记录！(源自历史文件: {src_file})", "ERROR")

                headers_wtd = list(df_source.columns)
                ws_wtd = fast_recreate_sheet(wb1, "问题单")
                ws_wtd.append(headers_wtd)
                for row_data in df_source.values.tolist():
                    ws_wtd.append(row_data)

            elif "支付绑定" in wb_name:
                try:
                    xls = pd.ExcelFile(file_path)
                    target_sheet = None
                    for sheet in xls.sheet_names:
                        if "绑定情况" in sheet or "绑定" in sheet:
                            target_sheet = sheet
                            break
                    if not target_sheet:
                        for sheet in xls.sheet_names:
                            df_temp = pd.read_excel(file_path, sheet_name=sheet, nrows=5)
                            if any("绑定" in str(c) for c in df_temp.columns):
                                target_sheet = sheet
                                break
                    if not target_sheet: target_sheet = xls.sheet_names[0]

                    df_bind_status = pd.read_excel(file_path, sheet_name=target_sheet, dtype=str)
                    df_bind_status.fillna("", inplace=True)
                    log(f"-> 🎯 成功读取带有绑定情况内容的子表：【{target_sheet}】！", "INFO")
                except Exception as e:
                    log(f"尝试读取【绑定情况】表失败，回退使用默认表: {e}", "WARN")
                    df_bind_status = df_source.copy()

                if len(df_bind_status.columns) > 0 and city in df_bind_status.iloc[:, 0].astype(str).unique():
                    df_source_filtered = df_bind_status[df_bind_status.iloc[:, 0] == city]
                else:
                    df_source_filtered = df_bind_status

                ws_bind = fast_recreate_sheet(wb1, "骑手支付绑定")

                headers_bind = list(df_source_filtered.columns)
                ws_bind.append(headers_bind)
                for row_data in df_source_filtered.values.tolist():
                    ws_bind.append(row_data)

            log(random.choice(theme['msg_success']).format(wb_name=wb_name))
            processed_count += 1
            current_progress += progress_step
            progress_callback(current_progress, f"狂奔中 {processed_count}/{total_files}: {wb_name[:12]}...")

        log(">>> 正在提提炼专属【日单量】汇总表，并匹配蓝橙单价...", "SYSTEM")
        progress_callback(0.85, "多维统计单价中...")
        ws_sht0 = wb1["配送单"]
        headers = []
        status_idx = -1
        fraud_idx = -1
        daily_summary = {}
        headers_abcd = ["A列", "B列", "C列", "D列", "日单量", "完成单", "不准时单", "是否欺诈单", "蓝橙单价"]

        rows_iterator = ws_sht0.iter_rows(values_only=True)
        try:
            headers = [str(c) if c is not None else "" for c in next(rows_iterator)]
            first_4_headers = headers[:4] + [""] * max(0, 4 - len(headers[:4]))
            headers_abcd = first_4_headers + ["日单量", "完成单", "不准时单", "是否欺诈单", "蓝橙单价"]

            try:
                status_idx = headers.index("运单状态")
            except ValueError:
                status_idx = 15
            try:
                fraud_idx = headers.index("是否欺诈单")
            except ValueError:
                fraud_idx = 17

            for row in rows_iterator:
                if len(row) > 0 and row[0] is not None and str(row[0]).strip() != "":
                    c0 = str(row[0]) if len(row) > 0 and row[0] is not None else ""
                    c1 = str(row[1]) if len(row) > 1 and row[1] is not None else ""
                    c2 = str(row[2]) if len(row) > 2 and row[2] is not None else ""
                    c3 = str(row[3]) if len(row) > 3 and row[3] is not None else ""
                    key = (c0, c1, c2, c3)

                    if key not in daily_summary: daily_summary[key] = {"total": 0, "completed": 0, "late": 0,
                                                                       "fraud": 0}
                    status_val = str(row[status_idx]).strip() if status_idx < len(row) and row[
                        status_idx] is not None else ""
                    fraud_val = str(row[fraud_idx]).strip() if fraud_idx < len(row) and row[
                        fraud_idx] is not None else ""

                    daily_summary[key]["total"] += 1
                    if status_val == "完成单-不准时":
                        daily_summary[key]["late"] += 1
                    elif status_val == "完成单":
                        daily_summary[key]["completed"] += 1
                    if fraud_val == "是": daily_summary[key]["fraud"] += 1
        except StopIteration:
            pass

        sorted_keys = sorted(daily_summary.keys(), key=lambda x: (str(x[0]), str(x[2])))

        if "日单量" in wb1.sheetnames:
            ws_daily = fast_recreate_sheet(wb1, "日单量")
        else:
            target_idx_daily = wb1.sheetnames.index("配送所得表") + 1 if "配送所得表" in wb1.sheetnames else 2
            ws_daily = wb1.create_sheet(title="日单量", index=target_idx_daily)

        ws_daily.append(headers_abcd)

        no_price_records = []
        rider_prices_dict = defaultdict(set)
        rider_price_dates_dict = defaultdict(lambda: defaultdict(list))

        for row_data in sorted_keys:
            summary = daily_summary[row_data]
            count_total = summary["total"]
            count_comp = summary["completed"]
            count_late = summary["late"]
            count_fraud = summary["fraud"]

            b_date = _norm_date(row_data[0])
            team_name = str(row_data[1]).strip()
            r_id = str(row_data[2]).replace('.0', '').strip()
            r_name = str(row_data[3]).strip()

            key_id = f"{b_date}_{team_name}_{r_id}"
            key_name = f"{b_date}_{team_name}_{r_name}"

            price = price_mapping_id.get(key_id)
            if not price: price = price_mapping_name.get(key_name)

            if price is not None and str(price).strip() != "":
                try:
                    price_float = float(price)
                    if price_float == 0.0:
                        price = "无单价"
                    else:
                        price = price_float
                except ValueError:
                    pass

            if not price or str(price).strip() == "": price = "无单价"

            full_row = list(row_data) + [count_total, count_comp, count_late, count_fraud, price]
            if price == "无单价": no_price_records.append(full_row)

            rider_key = r_id
            rider_prices_dict[rider_key].add(str(price))
            rider_price_dates_dict[rider_key][str(price)].append(b_date)

            ws_daily.append(full_row)

        def format_dates_streak(date_strs):
            if not date_strs: return ""
            d_objs = []
            for d in date_strs:
                try:
                    d_objs.append(datetime.strptime(d, '%Y-%m-%d'))
                except:
                    pass
            if not d_objs: return "；".join(date_strs)
            d_objs = sorted(list(set(d_objs)))
            streaks = []
            current_streak = [d_objs[0]]
            for i in range(1, len(d_objs)):
                if (d_objs[i] - current_streak[-1]).days == 1:
                    current_streak.append(d_objs[i])
                else:
                    streaks.append(current_streak)
                    current_streak = [d_objs[i]]
            streaks.append(current_streak)
            formatted = []
            for streak in streaks:
                start, end = streak[0], streak[-1]
                if start == end:
                    formatted.append(f"{start.day}日")
                else:
                    formatted.append(f"{start.day}日-{end.day}日")
            return "、".join(formatted)

        grouped = defaultdict(list)
        if no_price_records:
            log(f">>> 🚨 发现无单价数据，正在汇总并装进【蓝橙无单价明细】子表！", "WARN")
            if "蓝橙无单价明细" in wb1.sheetnames:
                ws_noprice = fast_recreate_sheet(wb1, "蓝橙无单价明细")
                current_idx = wb1.index(ws_noprice)
                if current_idx != 0: wb1.move_sheet(ws_noprice, offset=-current_idx)
            else:
                ws_noprice = wb1.create_sheet(title="蓝橙无单价明细", index=0)

            if "蓝橙无单价" in wb1.sheetnames:
                ws_old = fast_recreate_sheet(wb1, "蓝橙无单价")
                ws_old.sheet_state = 'hidden'

            ws_noprice.sheet_properties.tabColor = "FF9900"
            headers_noprice = ["团队名称", "骑手ID", "骑手姓名", "蓝橙无单价日期"]
            ws_noprice.append(headers_noprice)

            for rec in no_price_records:
                b_date = _norm_date(rec[0])
                team = str(rec[1]).strip()
                r_id = str(rec[2]).replace('.0', '').strip()
                r_name = str(rec[3]).strip()
                key = (team, r_id, r_name)
                grouped[key].append(b_date)

            for key, dates in grouped.items():
                date_str = format_dates_streak(dates) + "无单价"
                row_vals = [key[0], key[1], key[2], date_str]
                ws_noprice.append(row_vals)

        try:
            if "配送所得表" in wb1.sheetnames:
                ws_income_target = wb1["配送所得表"]
                history_scheme_map = {}
                month_str = last_date.month
                historical_merged_file = os.path.join(out_folder, f"{city}{month_str}月兼职已发.xlsx")

                if os.path.exists(historical_merged_file):
                    try:
                        df_hist = pd.read_excel(historical_merged_file, sheet_name="配送所得表", dtype=str)
                        t_col, i_col, s_col = None, None, None
                        for col_name in df_hist.columns:
                            c_str = str(col_name).strip()
                            if "团队" in c_str:
                                t_col = col_name
                            elif "id" in c_str.lower() or "编号" in c_str:
                                i_col = col_name
                            elif "薪资方案" in c_str:
                                s_col = col_name

                        if t_col is not None and i_col is not None and s_col is not None:
                            t_col_idx = df_hist.columns.get_loc(t_col)
                            i_col_idx = df_hist.columns.get_loc(i_col)
                            s_col_idx = df_hist.columns.get_loc(s_col)
                            for row in df_hist.itertuples(index=False):
                                t_val = str(row[t_col_idx]).replace('nan', '').strip()
                                i_val = str(row[i_col_idx]).replace('.0', '').replace('nan', '').strip()
                                s_val = str(row[s_col_idx]).replace('.0', '').replace('nan', '').strip()
                                if i_val and s_val and s_val != "None":
                                    key = i_val
                                    if key not in history_scheme_map: history_scheme_map[key] = set()
                                    history_scheme_map[key].add(s_val)
                    except Exception as e:
                        log(f"提取历史薪资方案出错: {e}", "WARN")

                remark_col_idx = -1
                leishen_price_col_idx = -1
                last_price_col_idx = -1

                for c in range(1, ws_income_target.max_column + 1):
                    cell_val_r3 = str(ws_income_target.cell(row=3, column=c).value or "").strip()
                    if "上期单价" in cell_val_r3:
                        last_price_col_idx = c
                        break

                for r in range(1, 6):
                    for c in range(1, ws_income_target.max_column + 1):
                        cell_val = str(ws_income_target.cell(row=r, column=c).value or "").strip()
                        if "备注" in cell_val:
                            remark_col_idx = c
                        elif "蓝橙单价" in cell_val:
                            leishen_price_col_idx = c

                noprice_lookup = defaultdict(list)
                for k, v in grouped.items():
                    r_id = k[1]
                    noprice_lookup[r_id].extend(v)
                noprice_lookup = {k: format_dates_streak(v) + "无单价" for k, v in noprice_lookup.items()}

                for row in ws_income_target.iter_rows(min_row=4):
                    team_val = str(row[1].value or "").strip()
                    id_val = str(row[2].value or "").replace('.0', '').strip()

                    if id_val:
                        lookup_key = id_val
                        remark_parts = []
                        if lookup_key in noprice_lookup: remark_parts.append(noprice_lookup[lookup_key])

                        if lookup_key in rider_prices_dict:
                            prices_set = rider_prices_dict[lookup_key]
                            if len(prices_set) > 1:
                                multi_price_remarks = []
                                for p, d_list in rider_price_dates_dict[lookup_key].items():
                                    if p == "无单价": continue
                                    streak_str = format_dates_streak(d_list)
                                    multi_price_remarks.append(f"{streak_str}单价{p}元")
                                if multi_price_remarks: remark_parts.append("；".join(multi_price_remarks))

                            if len(prices_set) > 1 and "无单价" in prices_set: prices_set.remove("无单价")

                            if leishen_price_col_idx != -1:
                                prices_list = sorted(list(prices_set))
                                if len(prices_list) == 1:
                                    final_leishen_price = to_numeric_if_possible(prices_list[0])
                                else:
                                    final_leishen_price = ";".join(prices_list)
                                row[leishen_price_col_idx - 1].value = final_leishen_price

                        if remark_col_idx != -1:
                            if remark_parts:
                                row[remark_col_idx - 1].value = "；".join(remark_parts)
                            else:
                                row[remark_col_idx - 1].value = ""  # 确保清空旧备注

                        if last_price_col_idx != -1:
                            if lookup_key in history_scheme_map:
                                schemes = history_scheme_map[lookup_key]
                                scheme_str = "；".join(sorted(list(schemes)))
                                row[last_price_col_idx - 1].value = scheme_str
                            else:
                                row[last_price_col_idx - 1].value = "无"

                if remark_col_idx != -1:
                    max_length = 0
                    col_letter = get_column_letter(remark_col_idx)
                    for row_val in ws_income_target.iter_rows(min_row=1, max_row=ws_income_target.max_row,
                                                              min_col=remark_col_idx, max_col=remark_col_idx,
                                                              values_only=True):
                        val = row_val[0]
                        if val is not None:
                            val_str = str(val)
                            val_len = len(val_str.encode('gbk', errors='ignore'))
                            if val_len > max_length: max_length = val_len
                    ws_income_target.column_dimensions[col_letter].width = max(11, min(max_length + 3.5, 60))

        except Exception as e:
            log(f"处理动态反写和高级备注合并时遭遇暗流: {e}", "WARN")

        if intercept_records:
            log(">>> 正在生成【拦截溯源】清单...", "SYSTEM")
            ws_intercept = wb1.create_sheet(title="拦截溯源")
            ws_intercept.sheet_properties.tabColor = "FF0000"
            ws_intercept.append(["拦截工作薄名称", "拦截类型", "运单号", "骑手ID", "骑手名称"])
            for rec in intercept_records:
                if len(rec) >= 5:
                    ws_intercept.append([rec[0], rec[1], rec[2], rec[3], rec[4]])
                else:
                    ws_intercept.append([rec[0], rec[1], rec[2], "", ""])

        log(">>> 正在为结果表钉上金装...", "SYSTEM")
        progress_callback(0.90, "精美装扮中...")
        border_style = Side(border_style="thin", color="000000")
        border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
        font_style = Font(name="微软雅黑", size=10)
        fill_style = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        red_font_strike = Font(name="微软雅黑", size=10, color="FF0000", bold=False, strike=True)
        red_font_header = Font(name="微软雅黑", size=10, color="FF0000", bold=False)
        red_border_style = Side(border_style="thin", color="FF0000")
        red_border = Border(left=red_border_style, right=red_border_style, top=red_border_style,
                            bottom=red_border_style)
        red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

        # 模板自带格式的表，绝对不要碰，以防破坏字体格式并节省海量时间
        template_sheets = ["配送所得表", "安全基金", "申请名单"]
        # 数据量可能很大的基础数据表
        raw_sheets = ["配送单", "违规索赔", "问题单", "骑手支付绑定"]

        for ws in wb1.worksheets:
            if ws.title in template_sheets:
                continue
                
            max_r = ws.max_row
            max_c = min(ws.max_column, 50)
            is_intercept = (ws.title == "拦截溯源")
            
            ws.freeze_panes = 'A2'

            # 仅动态图表或数据表首行做高速格式化
            iter_max_r = 1 if ws.title in raw_sheets else max_r
            
            if iter_max_r > 0:
                for row in ws.iter_rows(min_row=1, max_row=iter_max_r, min_col=1, max_col=max_c):
                    for cell in row:
                        val = cell.value
                        if val is None: continue
                        
                        cell.alignment = center_align
                        if is_intercept:
                            cell.border = red_border
                            cell.font = red_font_header if cell.row == 1 else red_font_strike
                        else:
                            cell.border = border
                            cell.font = font_style

                        if cell.row == 1:
                            cell.fill = fill_style

                        val_type = type(val)
                        if val_type is str:
                            val_str = val.strip()
                            if val_str.isdigit():
                                if len(val_str) < 15:
                                    cell.value = int(val_str)
                                    cell.number_format = '0'
                                else:
                                    cell.number_format = '@'
                            elif len(val_str) >= 15:
                                cell.number_format = '@'
                        elif val_type is float:
                            if cell.number_format not in ['yyyy/mm/dd', 'yyyy/m/d', 'm/d/yy', '@']:
                                if val.is_integer():
                                    cell.number_format = '0'
                                else:
                                    cell.number_format = '0.00'
                        elif val_type is int:
                            if cell.number_format not in ['yyyy/mm/dd', 'yyyy/m/d', 'm/d/yy', '@']: cell.number_format = '0'

            if ws.title != "配送所得表":
                max_lengths = [0] * max_c
                for row_val in ws.iter_rows(min_row=1, max_row=min(1000, max_r), min_col=1, max_col=max_c, values_only=True):
                    for idx, val in enumerate(row_val):
                        if val is not None:
                            val_str = str(val)
                            if not val_str.startswith('='):
                                val_len = len(val_str.encode('gbk', errors='ignore'))
                                if val_len > max_lengths[idx]: 
                                    max_lengths[idx] = val_len
                
                for col_idx, max_length in enumerate(max_lengths, 1):
                    col_letter = get_column_letter(col_idx)
                    ws.column_dimensions[col_letter].width = max(11, min(max_length + 3.5, 55))
                    
                for cell in ws[1]:
                    if cell.value is not None: cell.fill = red_fill if is_intercept else fill_style

        ws_income = wb1["配送所得表"]
        date_str = f"{first_date.strftime('%m.%d')}-{last_date.strftime('%m.%d')}"
        file_prefix = f"{city}{date_str}{selected_option}兼职"
        ws_income["A1"] = file_prefix

        ws_income.row_dimensions[1].height = 33
        ws_income.row_dimensions[3].height = 31

        for i, ws in enumerate(wb1.worksheets):
            try:
                if hasattr(ws, 'sheet_view'):
                    ws.sheet_view.tabSelected = (i == 0)
                elif hasattr(ws, 'views') and getattr(ws.views, 'sheetView', None):
                    ws.views.sheetView[0].tabSelected = (i == 0)
            except Exception:
                pass

        wb1.active = 0

        # =========================================================
        # 收尾：剥离专属表格，分别保存
        # =========================================================
        wb_special = openpyxl.Workbook()
        ws_default = wb_special.active
        has_special_data = False

        if "蓝橙无单价明细" in wb1.sheetnames:
            log(">>> 正在将【蓝橙无单价明细】打包剥离...", "SYSTEM")
            ws_noprice_old = wb1["蓝橙无单价明细"]
            ws_special_1 = ws_default
            ws_special_1.title = "蓝橙无单价明细"
            ws_special_1.sheet_properties.tabColor = "FF9900"
            has_special_data = True

            for row in ws_noprice_old.iter_rows(values_only=False):
                for cell in row:
                    new_cell = ws_special_1.cell(row=cell.row, column=cell.column, value=cell.value)
                    if cell.has_style:
                        new_cell.font = copy.copy(cell.font)
                        new_cell.border = copy.copy(cell.border)
                        new_cell.fill = copy.copy(cell.fill)
                        new_cell.alignment = copy.copy(cell.alignment)
                        new_cell.number_format = copy.copy(cell.number_format)

            for col_letter, col_dim in ws_noprice_old.column_dimensions.items(): ws_special_1.column_dimensions[
                col_letter].width = col_dim.width
            ws_special_1.freeze_panes = 'A2'
            del wb1["蓝橙无单价明细"]

        unbound_records = []
        if "安全基金" in wb1.sheetnames and "骑手支付绑定" in wb1.sheetnames:
            ws_fund = wb1["安全基金"]
            ws_bind = wb1["骑手支付绑定"]

            b_team_col, b_id_col, b_name_col, b_status_col = 2, 3, 4, 9
            for c in range(1, min(20, ws_bind.max_column + 1)):
                h_val = str(ws_bind.cell(row=1, column=c).value or "").strip()
                if "团队名称" in h_val:
                    b_team_col = c
                elif "id" in h_val.lower() or "编号" in h_val:
                    b_id_col = c
                elif "姓名" in h_val or "名称" in h_val:
                    b_name_col = c
                elif "绑" in h_val or "状态" in h_val:
                    b_status_col = c

            bind_map = {}
            for r in range(2, ws_bind.max_row + 1):
                b_team = str(ws_bind.cell(row=r, column=b_team_col).value or "").strip()
                b_id = str(ws_bind.cell(row=r, column=b_id_col).value or "").replace('.0', '').strip()
                b_name = str(ws_bind.cell(row=r, column=b_name_col).value or "").strip()
                b_status = str(ws_bind.cell(row=r, column=b_status_col).value or "").replace(" ", "").strip()
                if b_id: bind_map[b_id] = [b_team, b_id, b_name, b_status]

            for r in range(2, ws_fund.max_row + 1):
                f_team = str(ws_fund.cell(row=r, column=1).value or "").strip()
                f_id = str(ws_fund.cell(row=r, column=2).value or "").replace('.0', '').strip()
                f_name = str(ws_fund.cell(row=r, column=3).value or "").strip()

                if f_id:
                    if f_id in bind_map:
                        match_data = bind_map[f_id]
                        status_val = match_data[3]
                        if "未" in status_val or "否" in status_val or "失败" in status_val or status_val == "": unbound_records.append(
                            match_data)
                    else:
                        unbound_records.append([f_team, f_id, f_name, "不在支付绑定名单中"])

        log(f">>> 雷达探测结束，底层直接碰撞出 {len(unbound_records)} 条未绑记录！", "SYSTEM")

        if has_special_data:
            ws_special_2 = wb_special.create_sheet("发薪工具未绑名单")
        else:
            ws_special_2 = ws_default
            has_special_data = True

        ws_special_2.title = "发薪工具未绑名单"
        ws_special_2.sheet_properties.tabColor = "FF0000"

        write_row = 1
        headers_special2 = ["团队名称", "骑手ID", "骑手姓名", "绑定情况"]
        for c_idx, val in enumerate(headers_special2, 1): ws_special_2.cell(row=write_row, column=c_idx, value=val)
        for rec in unbound_records:
            ws_special_2.append(rec)

        border_style = Side(border_style="thin", color="000000")
        border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
        center_align = Alignment(horizontal="center", vertical="center")
        header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

        for row in ws_special_2.iter_rows(min_row=1, max_row=ws_special_2.max_row, min_col=1, max_col=4):
            for cell in row:
                cell.alignment = center_align
                cell.border = border
                if cell.row == 1:
                    cell.font = Font(name="微软雅黑", size=10)
                    cell.fill = header_fill
                else:
                    cell.font = Font(name="微软雅黑", size=10)
                    if cell.column == 2:
                        val_str = str(cell.value).strip()
                        if val_str.isdigit(): cell.value = int(val_str)
                        cell.number_format = '0'

        ws_special_2.column_dimensions['A'].width = 15
        ws_special_2.column_dimensions['B'].width = 15
        ws_special_2.column_dimensions['C'].width = 15
        ws_special_2.column_dimensions['D'].width = 30
        ws_special_2.freeze_panes = 'A2'

        if has_special_data:
            if "Sheet" in wb_special.sheetnames and len(wb_special.sheetnames) > 1: del wb_special["Sheet"]
            special_save_name = "蓝橙无单价&未绑名单.xlsx"
            special_save_path = os.path.join(out_folder, special_save_name)
            wb_special.save(special_save_path)

        save_name = f"{file_prefix}{datetime.now().strftime('%m%d')}.xlsx"
        final_save_path = os.path.join(out_folder, save_name)

        log(">>> 正在生成最终大包袱...", "SYSTEM")
        
        progress_callback(0.98, "生成最终报表中...")
        wb1.save(final_save_path)

        # 薪资公式防漏标红使用 openpyxl 是无效的（新写入公式的 data_only=None）
        # 改为由业务人员使用 Excel 自身能力核对，省下重复读取历史的 1~2 分钟！
        log(">>> 🎯 生成完毕！", "SUCCESS")

        if intercept_records: log(f">>> 🚨 红色警报：本次共拦截 {len(intercept_records)} 条重复数据！已悉数戴上红牌关押至【拦截溯源】子表！", "ERROR")

        elapsed_time = time.time() - start_time
        stats_info["elapsed_time"] = round(elapsed_time, 2)
        log(f">>> {random.choice(theme['msg_end'])}", "SYSTEM")
        log(f"本次运行消耗了: {elapsed_time:.2f} 秒")
        log(f"大功告成，产出文件: {final_save_path}", "INFO")
        progress_callback(1.00, "收工，准备下班~")
        finish_callback("success", final_save_path, stats_info)
    except Exception as e:
        log(f"哎呀！运行过程中出了个小意外: {str(e)}", "ERROR")
        progress_callback(0, "被未知力量打断")
        finish_callback("error", str(e), None)
