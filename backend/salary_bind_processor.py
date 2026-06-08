import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import os
import glob
import time
from datetime import datetime

async def run_salary_bind_gen(source_path, target_path=None):
    from tasks import create_progress_event, create_log_event, create_finish_event
    import asyncio
    
    yield create_progress_event(0.1)
    yield create_log_event(">>> 正在准备发薪支付数据...", "INFO")

    try:
        if not source_path or not os.path.exists(source_path):
            yield create_log_event(">>> [错误] 源目录不存在或未选择！", "ERROR")
            yield create_finish_event("error", "源目录不存在")
            return
            
        # default template and targets
        template_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "发薪支付数据-template.xlsx")
        if getattr(base_path_global, "value", None) and os.path.exists(os.path.join(base_path_global.value, "发薪支付数据-template.xlsx")):
            template_file = os.path.join(base_path_global.value, "发薪支付数据-template.xlsx")

        # In case we don't have the explicit template, we can create one dynamically
        # For a robust solution without template file dependencies:
        yield create_log_event(">>> 开始扫描源目录文件...", "INFO")
        files = glob.glob(os.path.join(source_path, "*.xlsx"))
        
        # Dictionaries to hold data for the sheets
        dict_jianzhi = []  # 兼职薪资数据
        dict_neitui = []   # 内推发放数据
        dict_ditui = []    # 地推发放数据
        dict_renyuan = []  # 人员信息
        dict_zhangdan = [] # 账单扣款明细

        header_jianzhi = None
        header_neitui = None
        header_ditui = None
        header_renyuan = None
        header_zhangdan = None
        
        yield create_progress_event(0.2)
        
        for i, file_path in enumerate(files):
            filename = os.path.basename(file_path)
            if "~$" in filename: continue
            yield create_log_event(f"-> 处理文件: {filename}", "INFO")
            try:
                # We want headers all as string, and values as they are
                df = pd.read_excel(file_path, dtype=str)
                df.columns = [str(c).strip() for c in df.columns]
                
                if "兼职薪资" in filename or "核算明细" in filename:
                    df_filtered = df[df.iloc[:, 5].notna() & (df.iloc[:, 5] != "")]
                    if header_jianzhi is None: header_jianzhi = list(df.columns)
                    dict_jianzhi.extend(df_filtered.values.tolist())
                    
                elif "merchantFreelancer" in filename:
                    df_filtered = df[df.iloc[:, 6].str.contains("内推", na=False)]
                    if header_neitui is None: header_neitui = list(df.columns)
                    dict_neitui.extend(df_filtered.values.tolist())
                    
                elif "项目人员信息" in filename:
                    df_filtered = df[df.iloc[:, 5].str.contains("地推", na=False)]
                    if header_ditui is None: header_ditui = list(df.columns)
                    dict_ditui.extend(df_filtered.values.tolist())
                    
                elif "薪资补发记录" in filename:
                    # drop Q:S, U:AD, AF:BJ => columns index mapping logic
                    if header_jianzhi is None: header_jianzhi = list(df.columns)
                    dict_jianzhi.extend(df.values.tolist())
                    
                elif "人员信息" in filename:
                    df_filtered = df[df.iloc[:, 0].notna() & (df.iloc[:, 0] != "")]
                    if header_renyuan is None: header_renyuan = list(df.columns)
                    dict_renyuan.extend(df_filtered.values.tolist())
                    
                elif "账单扣款" in filename or "结算" in filename:
                    df_filtered = df[df.iloc[:, 13].str.contains("账单", na=False)]
                    if header_zhangdan is None: header_zhangdan = list(df.columns)
                    dict_zhangdan.extend(df_filtered.values.tolist())
                    
            except Exception as e:
                yield create_log_event(f"读取文件出错 {filename}: {e}", "WARN")

        yield create_progress_event(0.5)
        
        # Process output
        mon = datetime.now().strftime("%m月")
        sh_jianzhi_name = f"{mon}兼职发薪数据"
        sh_weifa_name = f"{mon}未发清单"
        sh_zhangdan_name = f"{mon}账单扣款明细"

        df_renyuan = pd.DataFrame(dict_renyuan, columns=header_renyuan)
        df_jianzhi = pd.DataFrame(dict_jianzhi, columns=header_jianzhi)
        df_neitui = pd.DataFrame(dict_neitui, columns=header_neitui)
        df_ditui = pd.DataFrame(dict_ditui, columns=header_ditui)
        df_zhangdan = pd.DataFrame(dict_zhangdan, columns=header_zhangdan)

        # find unique people from renyuan (assume column 0 is identifier)
        unique_identifiers = []
        if not df_renyuan.empty:
            unique_identifiers = df_renyuan.iloc[:, 0].dropna().unique().tolist()
        else:
            # fallback to jianzhi column 0
            if not df_jianzhi.empty:
                unique_identifiers = df_jianzhi.iloc[:, 0].dropna().unique().tolist()

        # Create weifafang
        df_weifa = pd.DataFrame()
        if not df_jianzhi.empty and df_jianzhi.shape[1] >= 15:
            condition = df_jianzhi.iloc[:, 8].str.contains("未发", na=False) & \
                        (pd.to_numeric(df_jianzhi.iloc[:, 13], errors='coerce') > 0) & \
                        (df_jianzhi.iloc[:, 14] == "否")
            df_weifa = df_jianzhi[condition]
            
        yield create_progress_event(0.7)

        output_dir = os.path.join(os.path.dirname(source_path), f"(全量){mon}发薪支付数据" + datetime.now().strftime("%m%d"))
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)

        all_file_path = os.path.join(output_dir, f"(全量){mon}发薪支付数据{datetime.now().strftime('%m%d')}.xlsx")
        
        from openpyxl.utils.dataframe import dataframe_to_rows

        def beautify_workbook(workbook):
            font_style = Font(name="微软雅黑", size=10)
            center_align = Alignment(horizontal="center", vertical="center")
            header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            
            for ws in workbook.worksheets:
                for row in ws.iter_rows():
                    for cell in row:
                        cell.alignment = center_align
                        cell.font = font_style
                        if cell.row == 1:
                            cell.fill = header_fill
                ws.freeze_panes = 'A2'

        with pd.ExcelWriter(all_file_path, engine='openpyxl') as writer:
            if not df_jianzhi.empty: df_jianzhi.to_excel(writer, sheet_name=sh_jianzhi_name, index=False)
            if not df_weifa.empty: df_weifa.to_excel(writer, sheet_name=sh_weifa_name, index=False)
            if not df_zhangdan.empty: df_zhangdan.to_excel(writer, sheet_name=sh_zhangdan_name, index=False)
            if not df_neitui.empty: df_neitui.to_excel(writer, sheet_name="内推发放数据", index=False)
            if not df_ditui.empty: df_ditui.to_excel(writer, sheet_name="地推发放数据", index=False)
            if not df_renyuan.empty: df_renyuan.to_excel(writer, sheet_name="人员信息", index=False)
            
        wb_all = openpyxl.load_workbook(all_file_path)
        beautify_workbook(wb_all)
        wb_all.save(all_file_path)
        wb_all.close()

        yield create_progress_event(0.8)
        yield create_log_event(f"-> 全量汇总表已生成: {all_file_path}", "INFO")
        
        # Spilt part
        split_count = 0
        total_split = len(unique_identifiers)
        
        for identifier in unique_identifiers:
            split_count += 1
            if split_count % 5 == 0:
                yield create_progress_event(0.8 + 0.15 * (split_count / total_split))
                await asyncio.sleep(0.01)
                
            sp_jianzhi = df_jianzhi[df_jianzhi.iloc[:, 0] == str(identifier)] if not df_jianzhi.empty else pd.DataFrame()
            sp_weifa = df_weifa[df_weifa.iloc[:, 0] == str(identifier)] if not df_weifa.empty else pd.DataFrame()
            sp_zhangdan = df_zhangdan[df_zhangdan.iloc[:, 0] == str(identifier)] if not df_zhangdan.empty else pd.DataFrame()
            
            if sp_jianzhi.empty and sp_weifa.empty and sp_zhangdan.empty:
                continue
                
            sp_file_path = os.path.join(output_dir, f"{identifier}{mon}发薪支付数据{datetime.now().strftime('%m%d')}.xlsx")
            
            with pd.ExcelWriter(sp_file_path, engine='openpyxl') as writer:
                sp_jianzhi.to_excel(writer, sheet_name=sh_jianzhi_name, index=False)
                sp_weifa.to_excel(writer, sheet_name=sh_weifa_name, index=False)
                sp_zhangdan.to_excel(writer, sheet_name=sh_zhangdan_name, index=False)
                
            wb_sp = openpyxl.load_workbook(sp_file_path)
            beautify_workbook(wb_sp)
            wb_sp.save(sp_file_path)
            wb_sp.close()

        yield create_progress_event(1.0)
        yield create_log_event(f">>> 🎉 发薪工具绑定及拆分完毕！源文件数量：{len(files)}，拆分文件数量：{split_count}", "SUCCESS")
        yield create_finish_event("success", "绑定并拆分完成")
    except Exception as e:
        yield create_log_event(f">>> [错误] 执行发薪绑定合并时发生严重异常: {e}", "ERROR")
        yield create_finish_event("error", str(e))

class GlobalBase:
    pass
base_path_global = GlobalBase()
