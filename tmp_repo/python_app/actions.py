# /python_app/actions.py
import os
import glob
import re
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime

def run_remove_problem_orders(salary_file, problem_file, write_log):
    try:
        write_log(">>> 正在解析【问题单剔除工作薄】...", "INFO")
        xls = pd.ExcelFile(problem_file)
        all_problem_data = []

        kw_list = ["不准时单", "物流责", "超时", "t10", "投诉", "差评", "违规虚假", "不准时", "物流", "T10"]

        for sheet_name in xls.sheet_names:
            try:
                df = pd.read_excel(problem_file, sheet_name=sheet_name, dtype=str)
                df.columns = [str(c).strip() for c in df.columns]

                order_col = None
                if any(kw in sheet_name for kw in ["不准时单", "物流责", "超时", "t10", "不准时", "物流", "T10"]):
                    order_col = next((c for c in df.columns if "运单id" in c.lower()), None)
                elif any(kw in sheet_name for kw in ["投诉", "差评", "违规虚假"]):
                    order_col = next((c for c in df.columns if "运单号" in c), None)

                if not order_col:
                    order_col = next((c for c in df.columns if any(kw in c for kw in ["运单", "订单", "单号"])), None)

                is_remove_col = next((c for c in df.columns if "是否剔除" in c), None)
                reason_col = next((c for c in df.columns if "剔除原因" in c), None)

                if not order_col or not is_remove_col or not reason_col:
                    write_log(f"-> 跳过子表 [{sheet_name}]：缺少必要的(运单id/号)或(是否剔除/剔除原因)列。", "WARN")
                    continue

                type_name = sheet_name
                sheet_name_lower = sheet_name.lower()
                for kw in kw_list:
                    if kw.lower() in sheet_name_lower:
                        type_name = kw
                        break

                df = df[df[reason_col].notna() & (df[reason_col].str.strip() != "")]
                df = df[df[order_col].notna() & (df[order_col].str.strip() != "")]

                for _, row in df.iterrows():
                    val_id = str(row[order_col]).strip().replace(".0", "").lstrip("'")
                    val_is_rm = str(row[is_remove_col]).strip()
                    val_reason = str(row[reason_col]).strip()
                    all_problem_data.append({
                        "问题单类型": type_name,
                        "运单号": val_id,
                        "是否剔除": val_is_rm,
                        "剔除原因": val_reason
                    })
            except Exception as e:
                write_log(f"-> 解析子表 [{sheet_name}] 出错: {e}", "WARN")

        if not all_problem_data:
            write_log(">>> [警告] 未在问题单工作薄中找到任何有效的剔除记录，操作终止！", "WARN")
            return

        df_combined = pd.DataFrame(all_problem_data)

        match_dict = {}
        for _, row in df_combined.iterrows():
            key = f"{row['问题单类型']}_{row['运单号']}"
            match_dict[key] = {
                "是否剔除": row["是否剔除"],
                "剔除原因": row["剔除原因"]
            }

        write_log(f"-> 成功提取并合并了 {len(match_dict)} 条有效的问题单剔除规则！", "SUCCESS")
        write_log(">>> 正在打开【工资表工作薄】执行跨表匹配剔除...", "INFO")

        wb = openpyxl.load_workbook(salary_file)
        if "问题单" not in wb.sheetnames:
            write_log(">>> [错误] 工资表中未找到【问题单】子表，无法匹配！", "ERROR")
            return

        ws_wtd = wb["问题单"]

        type_col_idx, order_col_idx, is_rm_col_idx, reason_col_idx, deduct_col_idx = -1, -1, -1, -1, -1
        for c in range(1, ws_wtd.max_column + 1):
            val = str(ws_wtd.cell(row=1, column=c).value or "").strip()
            if val in ["类型", "问题单类型"]:
                type_col_idx = c
            elif any(kw in val for kw in ["运单", "订单", "单号"]):
                order_col_idx = c
            elif "是否剔除" in val:
                is_rm_col_idx = c
            elif "剔除原因" in val:
                reason_col_idx = c
            elif "扣款金额" in val:
                deduct_col_idx = c

        if -1 in [type_col_idx, order_col_idx, is_rm_col_idx, reason_col_idx]:
            write_log(">>> [错误] 工资表【问题单】子表缺少必要的匹配列！", "ERROR")
            return

        update_count = 0
        for r in range(2, ws_wtd.max_row + 1):
            c_type = str(ws_wtd.cell(row=r, column=type_col_idx).value or "").strip()
            c_order = str(ws_wtd.cell(row=r, column=order_col_idx).value or "").replace('.0', '').strip().lstrip("'")

            if c_type and c_order:
                key = f"{c_type}_{c_order}"
                if key in match_dict:
                    reason_val = match_dict[key]["剔除原因"]
                    ws_wtd.cell(row=r, column=is_rm_col_idx, value=match_dict[key]["是否剔除"])
                    ws_wtd.cell(row=r, column=reason_col_idx, value=reason_val)

                    if reason_val and deduct_col_idx != -1:
                        ws_wtd.cell(row=r, column=deduct_col_idx, value=0)

                    update_count += 1

        for col_idx in [type_col_idx, order_col_idx, is_rm_col_idx, reason_col_idx, deduct_col_idx]:
            if col_idx != -1:
                col_letter = get_column_letter(col_idx)
                max_len = 0
                for row_val in ws_wtd.iter_rows(min_row=1, max_row=ws_wtd.max_row, min_col=col_idx, max_col=col_idx, values_only=True):
                    if row_val[0] is not None:
                        val_len = len(str(row_val[0]).encode('gbk', errors='ignore'))
                        if val_len > max_len: max_len = val_len
                ws_wtd.column_dimensions[col_letter].width = max(11, min(max_len + 3.5, 55))

        archive_sheet_name = "剔除源头"
        if archive_sheet_name in wb.sheetnames:
            del wb[archive_sheet_name]
        ws_archive = wb.create_sheet(archive_sheet_name)
        ws_archive.sheet_properties.tabColor = "FF0000"

        headers = ["问题单类型", "运单号", "是否剔除", "剔除原因"]
        ws_archive.append(headers)
        for r_data in df_combined.itertuples(index=False):
            ws_archive.append(list(r_data))

        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

        border_style = Side(border_style="thin", color="000000")
        border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
        center_align = Alignment(horizontal="center", vertical="center")
        header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

        for col_idx in range(1, 5):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for row_idx in range(1, ws_archive.max_row + 1):
                cell = ws_archive.cell(row=row_idx, column=col_idx)
                cell.border = border
                cell.alignment = center_align
                if cell.row == 1:
                    cell.fill = header_fill
                    cell.font = Font(name="微软雅黑", size=10, bold=True)
                else:
                    cell.font = Font(name="微软雅黑", size=10)

                if cell.value is not None:
                    val_len = len(str(cell.value).encode('gbk', errors='ignore'))
                    if val_len > max_len: max_len = val_len

            ws_archive.column_dimensions[col_letter].width = max(12, min(max_len + 3.5, 50))

        ws_archive.freeze_panes = 'A2'

        wb.save(salary_file)
        write_log(f">>> 🎉 剔除完成！成功匹配并更新了 {update_count} 条记录。", "SUCCESS")
        write_log(f"-> 源头合集已归档至【{archive_sheet_name}】子表中，列宽已全部自适应排版！", "SUCCESS")

    except Exception as e:
        write_log(f">>> [错误] 执行问题单剔除时发生异常: {e}", "ERROR")

def run_raise_price(salary_file, price_file, write_log):
    from collections import defaultdict
    try:
        write_log(">>> 正在解析【蓝橙价格提审】文件...", "INFO")
        if price_file.lower().endswith('.csv'):
            try:
                df_price = pd.read_csv(price_file, encoding='gbk', dtype=str)
            except UnicodeDecodeError:
                df_price = pd.read_csv(price_file, encoding='utf-8', dtype=str)
        else:
            xls = pd.ExcelFile(price_file)
            sht_name = "兼职价格档案明细" if "兼职价格档案明细" in xls.sheet_names else xls.sheet_names[0]
            df_price = pd.read_excel(price_file, sheet_name=sht_name, dtype=str)

        for col in df_price.columns:
            df_price[col] = df_price[col].astype(str).str.replace(r'^[="\s\t]+|[="\s\t]+$', '', regex=True)

        rider_id_idx = next((i for i, c in enumerate(df_price.columns) if
                             ('id' in str(c).lower() or '编号' in str(c)) and ('骑手' in str(c) or '员工' in str(c))), 2)
        rider_name_idx = next((i for i, c in enumerate(df_price.columns) if
                               ('名' in str(c) or '姓名' in str(c)) and ('骑手' in str(c) or '员工' in str(c))), 3)

        def _norm_date(val):
            if pd.isna(val) or val is None: return ""
            s = str(val).strip().split(' ')[0].replace('/', '-')
            try:
                parts = s.split('-')
                if len(parts) >= 3: return f"{parts[0]}-{int(parts[1]):02d}-{int(parts[2]):02d}"
            except:
                pass
            return s

        price_mapping_id = {}
        price_mapping_name = {}

        for _, r in df_price.iterrows():
            try:
                if len(r) > 8:
                    date_val = _norm_date(r.iloc[1])
                    station_val = str(r.iloc[5]).strip()
                    price_val = str(r.iloc[8]).strip()
                    r_id = str(r.iloc[rider_id_idx]).replace('.0', '').strip() if rider_id_idx < len(r) else ""
                    r_name = str(r.iloc[rider_name_idx]).strip() if rider_name_idx < len(r) else ""
                    if r_id: price_mapping_id[f"{date_val}_{station_val}_{r_id}"] = price_val
                    if r_name: price_mapping_name[f"{date_val}_{station_val}_{r_name}"] = price_val
            except:
                pass

        write_log(f"-> 最新价格档案解析成功 (ID规则数: {len(price_mapping_id)} | 姓名规则数: {len(price_mapping_name)})", "INFO")
        write_log(">>> 正在打开【工资表工作薄】进行单价替换...", "INFO")

        wb = openpyxl.load_workbook(salary_file)
        if "日单量" not in wb.sheetnames or "配送所得表" not in wb.sheetnames:
            write_log(">>> [错误] 目标工资表缺少必要的【日单量】或【配送所得表】子表！", "ERROR")
            return

        ws_daily = wb["日单量"]
        date_col, team_col, id_col, name_col, price_col = 1, 2, 3, 4, 9

        rider_prices_dict = defaultdict(set)
        rider_price_dates_dict = defaultdict(lambda: defaultdict(list))
        no_price_records = []

        for r_idx in range(2, ws_daily.max_row + 1):
            b_date_raw = str(ws_daily.cell(row=r_idx, column=date_col).value or "").strip()
            if not b_date_raw or b_date_raw == "None": continue
            b_date = _norm_date(b_date_raw)
            team_name = str(ws_daily.cell(row=r_idx, column=team_col).value or "").strip()
            r_id = str(ws_daily.cell(row=r_idx, column=id_col).value or "").replace('.0', '').strip()
            r_name = str(ws_daily.cell(row=r_idx, column=name_col).value or "").strip()

            key_id = f"{b_date}_{team_name}_{r_id}"
            key_name = f"{b_date}_{team_name}_{r_name}"
            price = price_mapping_id.get(key_id)
            if not price: price = price_mapping_name.get(key_name)

            if price is not None and str(price).strip() != "" and str(price).lower() != "nan":
                try:
                    price_float = float(price)
                    if price_float == 0.0:
                        price = "无单价"
                    else:
                        price = price_float
                except ValueError:
                    pass
            else:
                price = "无单价"

            if not price or str(price).strip() == "" or str(price).lower() == "nan": price = "无单价"

            ws_daily.cell(row=r_idx, column=price_col, value=price)
            rider_key = f"{team_name}_{r_id}"
            rider_prices_dict[rider_key].add(str(price))
            rider_price_dates_dict[rider_key][str(price)].append(b_date)
            if price == "无单价": no_price_records.append((b_date, team_name, r_id, r_name))

        def format_dates_streak(date_strs):
            if not date_strs: return ""
            d_objs = []
            for d in date_strs:
                try:
                    d_objs.append(datetime.strptime(d, '%Y-%m-%d'))
                except:
                    pass
            if not d_objs: return "；".join(date_strs)
            d_objs = sorted(list(set(d_objs)))
            streaks = []
            current_streak = [d_objs[0]]
            for i in range(1, len(d_objs)):
                if (d_objs[i] - current_streak[-1]).days == 1:
                    current_streak.append(d_objs[i])
                else:
                    streaks.append(current_streak)
                    current_streak = [d_objs[i]]
            streaks.append(current_streak)
            formatted = []
            for streak in streaks:
                start, end = streak[0], streak[-1]
                if start == end:
                    formatted.append(f"{start.day}日")
                else:
                    formatted.append(f"{start.day}日-{end.day}日")
            return "、".join(formatted)

        def to_numeric_if_possible(val_str):
            if not val_str: return val_str
            try:
                f = float(val_str)
                if f.is_integer(): return int(f)
                return f
            except ValueError:
                return val_str

        grouped = defaultdict(list)
        for rec in no_price_records: grouped[(rec[1], rec[2], rec[3])].append(rec[0])

        noprice_lookup = {f"{k[0]}_{k[1]}": format_dates_streak(v) + "无单价" for k, v in grouped.items()}

        ws_income = wb["配送所得表"]
        remark_col_idx = -1
        leishen_price_col_idx = -1

        for r in range(1, 6):
            for c in range(1, ws_income.max_column + 1):
                cell_val = str(ws_income.cell(row=r, column=c).value or "").strip()
                if "备注" in cell_val:
                    remark_col_idx = c
                elif "蓝橙单价" in cell_val:
                    leishen_price_col_idx = c

        for r_idx in range(4, ws_income.max_row + 1):
            team_val = str(ws_income.cell(row=r_idx, column=2).value or "").strip()
            id_val = str(ws_income.cell(row=r_idx, column=3).value or "").replace('.0', '').strip()

            if team_val or id_val:
                lookup_key = f"{team_val}_{id_val}"
                remark_parts = []

                if lookup_key in noprice_lookup: remark_parts.append(noprice_lookup[lookup_key])

                if lookup_key in rider_prices_dict:
                    prices_set = rider_prices_dict[lookup_key]
                    if len(prices_set) > 1:
                        multi_price_remarks = []
                        for p, d_list in rider_price_dates_dict[lookup_key].items():
                            if p == "无单价": continue
                            streak_str = format_dates_streak(d_list)
                            multi_price_remarks.append(f"{streak_str}单价{p}元")
                        if multi_price_remarks: remark_parts.append("；".join(multi_price_remarks))

                    if "无单价" in prices_set: prices_set.remove("无单价")

                    if leishen_price_col_idx != -1:
                        if not prices_set:
                            final_leishen_price = "无单价"
                        else:
                            prices_list = sorted(list(prices_set))
                            if len(prices_list) == 1:
                                final_leishen_price = to_numeric_if_possible(prices_list[0])
                            else:
                                final_leishen_price = ";".join(prices_list)
                        ws_income.cell(row=r_idx, column=leishen_price_col_idx, value=final_leishen_price)
                else:
                    if leishen_price_col_idx != -1:
                        ws_income.cell(row=r_idx, column=leishen_price_col_idx, value="无单价")

                if remark_col_idx != -1:
                    if remark_parts:
                        ws_income.cell(row=r_idx, column=remark_col_idx, value="；".join(remark_parts))
                    else:
                        ws_income.cell(row=r_idx, column=remark_col_idx, value="")

        wb.save(salary_file)

        write_log(">>> 正在唤醒底层 Excel 引擎，执行薪资方案比对及精准标红...", "SYSTEM")
        try:
            import win32com.client
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False

            abs_final_path = os.path.abspath(salary_file)
            wb_com = excel.Workbooks.Open(abs_final_path, UpdateLinks=False)
            ws_com = wb_com.Sheets("配送所得表")

            max_row = ws_com.UsedRange.Rows.Count + ws_com.UsedRange.Row - 1
            max_col = ws_com.UsedRange.Columns.Count + ws_com.UsedRange.Column - 1

            lei_col = -1
            plan_col = -1

            header_vals = ws_com.Range(ws_com.Cells(1, 1), ws_com.Cells(5, max_col)).Value
            if header_vals:
                for r in range(min(5, len(header_vals))):
                    for c in range(len(header_vals[r])):
                        val = str(header_vals[r][c] or "").strip()
                        if "蓝橙单价" in val: lei_col = c + 1
                        if "薪资方案" in val: plan_col = c + 1

            if lei_col != -1 and plan_col != -1 and max_row >= 4:
                def is_equal(v1, v2):
                    s1 = str(v1).strip() if v1 is not None else ""
                    s2 = str(v2).strip() if v2 is not None else ""
                    if s1 == s2: return True
                    try:
                        return float(s1) == float(s2)
                    except:
                        return False

                range_lei = ws_com.Range(ws_com.Cells(4, lei_col), ws_com.Cells(max_row, lei_col))
                range_plan = ws_com.Range(ws_com.Cells(4, plan_col), ws_com.Cells(max_row, plan_col))

                range_plan.Interior.ColorIndex = -4142

                lei_values = range_lei.Value
                plan_values = range_plan.Value

                range_plan.Value = plan_values

                if max_row == 4:
                    lei_values = [[lei_values]]
                    plan_values = [[plan_values]]

                for i in range(len(lei_values)):
                    val_lei = lei_values[i][0]
                    val_plan = plan_values[i][0]
                    str_lei = str(val_lei).strip() if val_lei is not None else ""
                    str_plan = str(val_plan).strip() if val_plan is not None else ""

                    if str_lei and str_plan and str_lei != "无单价" and not is_equal(val_lei, val_plan):
                        ws_com.Cells(i + 4, plan_col).Interior.Color = 255  # 标红

            wb_com.Save()
            wb_com.Close()
            excel.Quit()
            write_log(">>> 🎯 薪资方案比对完成，差异项已无情标红！", "SUCCESS")
        except ImportError:
            write_log("系统缺少 pywin32 库，跳过单价标红操作！", "WARN")
        except Exception as e:
            write_log(f"COM 引擎比对标红时发生异常: {e}", "WARN")
            try:
                excel.Quit()
            except:
                pass

        write_log(f">>> 🎉 蓝橙单价重刷全部完成！新文件已覆盖: {os.path.basename(salary_file)}", "SUCCESS")
    except Exception as e:
        write_log(f">>> [错误] 重刷蓝橙单价时发生异常: {e}", "ERROR")

def run_summary_parttime(folder, city, write_log, open_file):
    try:
        merged_inc_data = []
        merged_fun_data = []
        inc_headers = []
        fun_headers = []

        write_log(">>> 正在极速提取目录下已有工作簿进行合并...", "SYSTEM")

        for f_path in glob.glob(os.path.join(folder, "*.xlsx")):
            f_name = os.path.basename(f_path)
            if "~$" in f_name or "(重复)" in f_name or "蓝橙无单价" in f_name or "兼职已发汇总" in f_name: continue

            wb_name_no_ext = os.path.splitext(f_name)[0]

            try:
                wb = openpyxl.load_workbook(f_path, data_only=True, read_only=True)
            except Exception as e:
                write_log(f"跳过不可读文件: {f_name} - {str(e)}", "WARN")
                continue

            try:
                if "安全基金" in wb.sheetnames:
                    ws_fun = wb["安全基金"]
                    rows = list(ws_fun.iter_rows(values_only=True))
                    if len(rows) >= 1:
                        if not fun_headers:
                            h_vals = rows[0]
                            if h_vals: fun_headers = ["工作薄名称"] + [str(x) if x is not None else "" for x in h_vals]

                        if len(rows) >= 2:
                            for row_data in rows[1:]:
                                if len(row_data) > 1 and (row_data[0] or row_data[1]): 
                                    merged_fun_data.append([wb_name_no_ext] + list(row_data))
            except Exception:
                pass

            try:
                if "配送所得表" in wb.sheetnames:
                    ws_inc = wb["配送所得表"]
                    rows = list(ws_inc.iter_rows(values_only=True))
                    if len(rows) >= 3:
                        if not inc_headers:
                            h_vals = rows[2]
                            if h_vals: inc_headers = ["工作薄名称"] + [str(x) if x is not None else "" for x in h_vals]

                        if len(rows) >= 4:
                            for row_data in rows[3:]:
                                if len(row_data) > 2 and (row_data[1] or row_data[2]): 
                                    merged_inc_data.append([wb_name_no_ext] + list(row_data))
            except Exception:
                pass

            try: wb.close()
            except Exception: pass

        if merged_inc_data or merged_fun_data:
            merged_filename = f"{city}兼职已发汇总.xlsx"
            merged_filepath = os.path.join(folder, merged_filename)

            def clean_and_type_data(data, headers):
                safe_headers = []
                seen = set()
                for i, h in enumerate(headers):
                    h_str = str(h) if h is not None else f"Unnamed_{i}"
                    h_str = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', h_str).strip()
                    if not h_str: h_str = f"Unnamed_{i}"

                    new_h = h_str
                    count = 1
                    while new_h in seen:
                        new_h = f"{h_str}_{count}"
                        count += 1
                    seen.add(new_h)
                    safe_headers.append(new_h)

                max_len = len(safe_headers)
                cleaned_data = []

                for row in data:
                    new_row = []
                    row_iter = list(row)[:max_len]
                    row_iter += [None] * (max_len - len(row_iter))

                    for val in row_iter:
                        if val is None:
                            new_row.append("")
                            continue
                        if hasattr(val, 'strftime'):
                            try:
                                val = val.strftime("%Y-%m-%d %H:%M:%S")
                            except:
                                val = str(val)

                        val_str = str(val).strip()
                        val_str = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', val_str)

                        if val_str.startswith('='): val_str = "'" + val_str
                        if val_str.endswith('.0'): val_str = val_str[:-2]
                        if val_str.lower() in ['nan', 'none', 'inf', '-inf', 'nat']:
                            new_row.append("")
                            continue

                        if not val_str:
                            new_row.append("")
                        elif val_str.isdigit():
                            if len(val_str) < 15:
                                new_row.append(int(val_str))
                            else:
                                new_row.append(str(val_str))
                        else:
                            try:
                                f_val = float(val_str)
                                new_row.append(f_val)
                            except ValueError:
                                new_row.append(val_str)
                    cleaned_data.append(new_row)
                return pd.DataFrame(cleaned_data, columns=safe_headers)

            with pd.ExcelWriter(merged_filepath, engine='openpyxl') as writer:
                if merged_inc_data:
                    df_inc = clean_and_type_data(merged_inc_data, inc_headers)
                    df_inc.to_excel(writer, sheet_name="配送所得表", index=False)
                if merged_fun_data:
                    df_fun = clean_and_type_data(merged_fun_data, fun_headers)
                    df_fun.to_excel(writer, sheet_name="安全基金", index=False)

                from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
                from openpyxl.utils import get_column_letter

                border_style = Side(border_style="thin", color="000000")
                border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
                font_style = Font(name="微软雅黑", size=10)
                center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
                header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

                for sheet_name in writer.sheets:
                    ws = writer.sheets[sheet_name]
                    max_r = ws.max_row
                    max_c = ws.max_column

                    ws.freeze_panes = 'A2'
                    for row_idx in range(1, max_r + 1): ws.row_dimensions[row_idx].height = 19.5

                    for row in ws.iter_rows(min_row=1, max_row=max_r, min_col=1, max_col=max_c):
                        for cell in row:
                            cell.alignment = center_align
                            cell.border = border
                            cell.font = font_style

                            if cell.row == 1:
                                cell.fill = header_fill
                            else:
                                if sheet_name == "配送所得表":
                                    val_str = str(cell.value).strip() if cell.value is not None else ""
                                    if val_str.isdigit() and len(val_str) >= 10:
                                        cell.number_format = '0'
                                    elif isinstance(cell.value, (int, float)) and cell.value > 1000000000:
                                        cell.number_format = '0'
                                    else:
                                        cell.number_format = 'General'
                                else:
                                    if type(cell.value) is int:
                                        cell.number_format = '0'
                                    elif type(cell.value) is float:
                                        cell.number_format = '0.00'
                                    elif type(cell.value) is str and str(cell.value).isdigit() and len(str(cell.value)) >= 15:
                                        cell.number_format = '@'

                    for col_idx in range(1, max_c + 1):
                        col_letter = get_column_letter(col_idx)
                        max_length = 0
                        for row_val in ws.iter_rows(min_row=1, max_row=min(500, max_r), min_col=col_idx, max_col=col_idx, values_only=True):
                            val = row_val[0]
                            if val is not None:
                                val_str = str(val)
                                val_len = len(val_str.encode('gbk', errors='ignore'))
                                if val_len > max_length: max_length = val_len
                        ws.column_dimensions[col_letter].width = max(11, min(max_length + 3.5, 55))

            write_log(f">>> 🎉 成功合并文件，并完成金装美化，存至: {merged_filepath}", "SUCCESS")
            open_file(merged_filepath)
        else:
            write_log(">>> 所选目录为空或只有无关文件，未生成合并总表~", "WARN")

    except ImportError:
        write_log("系统缺少 pywin32 库，无法合并文件！", "ERROR")
    except Exception as e:
        write_log(f"合并报表时出现意外: {e}", "ERROR")
        try:
            excel.Quit()
        except:
            pass
