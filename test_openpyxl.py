import openpyxl

wb = openpyxl.Workbook()
ws1 = wb.active
ws1.title = "Sheet1"
ws1['A1'] = 10

ws2 = wb.create_sheet(title="Sheet2")
ws2['A1'] = "=Sheet1!A1"

idx = wb.sheetnames.index("Sheet1")
del wb["Sheet1"]
ws3 = wb.create_sheet(title="Sheet1", index=idx)
ws3['A1'] = 20

wb.save("test_openpyxl.xlsx")

wb2 = openpyxl.load_workbook("test_openpyxl.xlsx")
print(wb2["Sheet2"]['A1'].value)
